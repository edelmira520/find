#!/usr/bin/env python3
import argparse
import copy
import html
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

SHEET_NAME = "樊登读书APP每周书籍"
TITLE_COL = 3
STATUS_COL = 13
COVER_COLUMNS = {
    9: ("cover", "书封"),
    10: ("standing", "立封"),
}
REMOVED_STATUSES = {"下架", "已下架", "停售", "不上架", "停用", "removed", "discontinued", "inactive", "offline"}
DEFAULT_SPEAKER_ID = "fandeng"
DEFAULT_SPEAKER_NAME = "樊登"


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_speaker_args(speaker_id, speaker_name):
    speaker_id = clean_text(speaker_id) or DEFAULT_SPEAKER_ID
    speaker_name = clean_text(speaker_name) or DEFAULT_SPEAKER_NAME
    if speaker_id == "all" or speaker_name == "全部":
        raise ValueError("“全部”只是前端筛选项，导入时必须指定真实讲书人")
    return speaker_id, speaker_name


def clean_title(value):
    title = clean_text(value)
    title = re.sub(r"^[《<〈\s]+|[》>〉\s]+$", "", title)
    return re.sub(r"\s+", " ", title).strip()


def normalize_status(value):
    return clean_text(value).lower()


def should_keep_status(value):
    return normalize_status(value) not in REMOVED_STATUSES


def slug_id(index):
    return f"book_{index:03d}"


def ensure_speaker(data_dir, speaker_id, speaker_name):
    data_dir.mkdir(parents=True, exist_ok=True)
    speakers_path = data_dir / "speakers.json"
    speakers = []
    if speakers_path.exists():
        try:
            speakers = json.loads(speakers_path.read_text(encoding="utf-8"))
            if not isinstance(speakers, list):
                speakers = []
        except json.JSONDecodeError:
            speakers = []
    if not any(speaker.get("id") == speaker_id for speaker in speakers if isinstance(speaker, dict)):
        speakers.append({"id": speaker_id, "name": speaker_name})
    speakers_path.write_text(json.dumps(speakers, ensure_ascii=False, indent=2), encoding="utf-8")


def image_extension(image):
    path = getattr(image, "path", "") or ""
    suffix = Path(path).suffix.lower()
    if suffix in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
        return ".jpg" if suffix == ".jpeg" else suffix
    fmt = (getattr(image, "format", "") or "").lower()
    if fmt in ["jpeg", "jpg"]:
        return ".jpg"
    if fmt in ["png", "webp", "gif"]:
        return f".{fmt}"
    return ".jpg"


def image_anchor(image):
    anchor = getattr(image, "anchor", None)
    if not anchor:
        return None
    try:
        return anchor._from.row + 1, anchor._from.col + 1
    except Exception:
        return None


def extract_image_bytes(image):
    image_copy = copy.copy(image)
    return image_copy._data()


def write_image(image, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(extract_image_bytes(image))


def resolve_actual_version(book):
    covers = book["covers"]
    return "cover" if covers.get("cover") else "noCover"


def rel_img(path):
    if not path:
        return ""
    return path.replace("\\", "/")


def preview_card(book, duplicate_titles):
    actual = resolve_actual_version(book)
    covers = book["covers"]
    cells = []
    columns = [
        ("书封", covers.get("cover", "")),
        ("立封", covers.get("standing", "")),
    ]
    for label, path in columns:
        if path:
            img = f'<img src="{html.escape(path)}" alt="{html.escape(label)}">'
        else:
            img = '<div class="missing">缺失</div>'
        cells.append(f"<div class='cover-cell'><div class='label'>{html.escape(label)}</div>{img}</div>")

    duplicate = book["title"] in duplicate_titles
    duplicate_badge = '<strong class="dup-badge">重复书名</strong>' if duplicate else ""
    filters = ["all", actual]
    if actual == "noCover":
        filters.append("missing")
    if duplicate:
        filters.append("duplicate")

    return f"""
    <article class="book-card {html.escape(actual)}" data-filters="{' '.join(filters)}">
      <header>
        <h2>{html.escape(book["title"])}{duplicate_badge}</h2>
        <span class="actual">当前展示：{version_label(actual)}</span>
      </header>
      <div class="covers">{''.join(cells)}</div>
    </article>
    """


def version_label(version):
    return {
        "cover": "书封",
        "noCover": "缺封面",
    }.get(version, version)


def build_preview_html(books, report):
    duplicate_titles = {item["title"] for item in report.get("duplicateTitles", [])}
    cards = "\n".join(preview_card(book, duplicate_titles) for book in books)
    summary = html.escape(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>书封导入核对</title>
  <style>
    :root {{
      color-scheme: light;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: #f3efe7;
      color: #25221d;
    }}
    body {{
      margin: 0;
      padding: 24px;
      background: repeating-linear-gradient(0deg, rgba(117, 94, 62, 0.035) 0 1px, transparent 1px 32px);
    }}
    h1 {{
      margin: 0 0 12px;
      font-size: 24px;
    }}
    .summary {{
      white-space: pre-wrap;
      background: #fffdf8;
      border: 1px solid #e2d8ca;
      border-radius: 8px;
      padding: 16px;
      margin-bottom: 12px;
    }}
    .filters {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin: 0 0 18px;
    }}
    button {{
      border: 1px solid #e2d8ca;
      border-radius: 999px;
      background: #fffdf8;
      padding: 8px 12px;
      cursor: pointer;
    }}
    button.active {{
      border-color: #116a5b;
      color: #116a5b;
      background: #edf8f5;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(390px, 1fr));
      gap: 16px;
    }}
    .book-card {{
      background: #fffdf8;
      border: 1px solid #e2d8ca;
      border-radius: 8px;
      padding: 14px;
      box-shadow: 0 12px 28px rgba(78, 58, 34, 0.1);
    }}
    .book-card header {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 12px;
    }}
    .book-card h2 {{
      margin: 0;
      font-size: 17px;
      line-height: 1.35;
    }}
    .book-card .actual {{
      flex: 0 0 auto;
      font-size: 12px;
      color: #51483c;
      background: #f7f1e8;
      border-radius: 999px;
      padding: 4px 8px;
    }}
    .covers {{
      display: grid;
      grid-template-columns: repeat(2, 1fr);
      gap: 10px;
    }}
    .cover-cell {{
      min-width: 0;
    }}
    .label {{
      font-size: 12px;
      color: #746f66;
      margin-bottom: 6px;
    }}
    img {{
      display: block;
      width: 100%;
      aspect-ratio: 2 / 3;
      object-fit: contain;
      background: #f5ecdf;
      border: 1px solid #e2d8ca;
      border-radius: 6px;
    }}
    .missing {{
      display: grid;
      place-items: center;
      width: 100%;
      aspect-ratio: 2 / 3;
      background: #f5ecdf;
      border: 1px dashed #c9bba8;
      border-radius: 6px;
      color: #9a8d7b;
      font-size: 13px;
    }}
    .dup-badge {{
      display: inline-flex;
      margin-left: 8px;
      padding: 3px 7px;
      border-radius: 999px;
      color: #b42318;
      background: #fff1ee;
      font-size: 12px;
      vertical-align: middle;
    }}
    .cover {{ border-left: 4px solid #0f766e; }}
    .noCover {{ border-left: 4px solid #b42318; }}
    .book-card.hidden {{ display: none; }}
  </style>
</head>
<body>
  <h1>书封导入核对</h1>
  <div class="summary">{summary}</div>
  <div class="filters">
    <button class="active" data-filter="all">全部</button>
    <button data-filter="missing">缺封面</button>
    <button data-filter="duplicate">重复书名</button>
    <button data-filter="cover">书封</button>
  </div>
  <main class="grid">{cards}</main>
  <script>
    const buttons = Array.from(document.querySelectorAll("button[data-filter]"));
    const cards = Array.from(document.querySelectorAll(".book-card"));
    buttons.forEach(button => button.addEventListener("click", () => {{
      buttons.forEach(item => item.classList.toggle("active", item === button));
      const filter = button.dataset.filter;
      cards.forEach(card => {{
        card.classList.toggle("hidden", !card.dataset.filters.split(" ").includes(filter));
      }});
    }}));
  </script>
</body>
</html>
"""


def convert(excel_path, output_dir, speaker_id=DEFAULT_SPEAKER_ID, speaker_name=DEFAULT_SPEAKER_NAME):
    from openpyxl import load_workbook

    output_dir = Path(output_dir)
    data_dir = output_dir
    speaker_id, speaker_name = normalize_speaker_args(speaker_id, speaker_name)
    ensure_speaker(data_dir, speaker_id, speaker_name)

    covers_dir = data_dir / "covers"
    if covers_dir.exists():
        shutil.rmtree(covers_dir)
    covers_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(excel_path, read_only=False, data_only=True)
    sheet_warning = ""
    if SHEET_NAME in workbook.sheetnames:
        worksheet = workbook[SHEET_NAME]
    else:
        worksheet = workbook.worksheets[0]
        sheet_warning = f"警告：未找到指定 sheet「{SHEET_NAME}」，实际使用第一个 sheet「{worksheet.title}」。"

    images_by_cell = defaultdict(list)
    ignored_images = []
    for index, image in enumerate(getattr(worksheet, "_images", []), 1):
        anchor = image_anchor(image)
        if not anchor:
            ignored_images.append({"imageIndex": index, "reason": "missing anchor"})
            continue
        images_by_cell[anchor].append(image)

    books = []
    row_reports = []
    blank_title_rows = []
    skipped_empty_rows = []
    removed_status_rows = []
    today = date.today().isoformat()

    for row in range(2, worksheet.max_row + 1):
        title = clean_title(worksheet.cell(row, TITLE_COL).value)
        if not title:
            has_content = any(clean_text(worksheet.cell(row, col).value) for col in range(1, 15))
            if has_content:
                blank_title_rows.append(row)
            else:
                skipped_empty_rows.append(row)
            continue

        note = clean_text(worksheet.cell(row, STATUS_COL).value)
        if not should_keep_status(note):
            removed_status_rows.append(row)
            continue

        book_id = slug_id(len(books) + 1)
        book = {
            "id": book_id,
            "title": title,
            "speakerId": speaker_id,
            "speakerName": speaker_name,
            "note": note,
            "status": "active",
            "covers": {
                "cover": "",
                "standing": "",
            },
            "createdAt": today,
            "updatedAt": today,
        }

        extracted = {}
        multiple_images = []
        for col, (slot, label) in COVER_COLUMNS.items():
            images = images_by_cell.get((row, col), [])
            if not images:
                continue
            if len(images) > 1:
                multiple_images.append({"column": col, "count": len(images)})
            image = images[0]
            extension = image_extension(image)
            filename = f"{book_id}_{slot}{extension}"
            output_path = covers_dir / filename
            write_image(image, output_path)
            book["covers"][slot] = rel_img(f"covers/{filename}")
            extracted[slot] = rel_img(f"covers/{filename}")

        books.append(book)
        row_reports.append({
            "row": row,
            "id": book_id,
            "title": title,
            "note": book["note"],
            "status": book["status"],
            "actualVersion": resolve_actual_version(book),
            "covers": extracted,
            "multipleImagesInCell": multiple_images,
        })

    title_counts = Counter(book["title"] for book in books)
    duplicate_titles = [
        {"title": title, "count": count, "ids": [book["id"] for book in books if book["title"] == title]}
        for title, count in title_counts.items()
        if count > 1
    ]
    missing_covers = [
        {"id": book["id"], "title": book["title"]}
        for book in books
        if resolve_actual_version(book) == "noCover"
    ]
    actual_counts = Counter(resolve_actual_version(book) for book in books)

    report = {
        "summary": {
            "source": str(excel_path),
            "requestedSheet": SHEET_NAME,
            "usedSheet": worksheet.title,
            "sheetWarning": sheet_warning,
            "generatedAt": today,
            "columnMapping": {
                "title": "C列：书籍名称",
                "cover": "I列：书封",
                "standing": "J列：立封",
                "note": "M列：备注",
            },
            "importedCount": len(books),
            "missingCoverCount": len(missing_covers),
            "duplicateTitleCount": len(duplicate_titles),
            "removedStatusRowCount": len(removed_status_rows),
            "skippedEmptyRowCount": len(skipped_empty_rows),
            "blankTitleWithContentRowCount": len(blank_title_rows),
            "actualVersionCounts": dict(actual_counts),
        },
        "blankTitleRows": blank_title_rows,
        "skippedEmptyRows": skipped_empty_rows,
        "removedStatusRows": removed_status_rows,
        "duplicateTitles": duplicate_titles,
        "missingCovers": missing_covers,
        "ignoredImages": ignored_images,
        "rows": row_reports,
    }

    (data_dir / "books.json").write_text(json.dumps(books, ensure_ascii=False, indent=2), encoding="utf-8")
    (data_dir / "import-report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    (data_dir / "import-preview.html").write_text(build_preview_html(books, report), encoding="utf-8")

    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser(description="Import book covers from Excel embedded images.")
    parser.add_argument("excel", help="Path to the source .xlsx file")
    parser.add_argument("--out", default="data", help="Output data directory")
    parser.add_argument("--speaker-id", default=DEFAULT_SPEAKER_ID, help="Lecturer/speaker id for imported books")
    parser.add_argument("--speaker-name", default=DEFAULT_SPEAKER_NAME, help="Lecturer/speaker name for imported books")
    args = parser.parse_args()
    convert(Path(args.excel), Path(args.out), speaker_id=args.speaker_id, speaker_name=args.speaker_name)


if __name__ == "__main__":
    main()
