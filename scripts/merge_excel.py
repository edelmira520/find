#!/usr/bin/env python3
import argparse
import copy
import html
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path


SHEET_NAME = "樊登读书APP每周书籍"
TITLE_COL = 3
STATUS_COL = 13
COVER_COLUMNS = {
    11: ("custom", "flat", "custom_flat", "自制平封"),
    9: ("original", "flat", "original_flat", "原版平封"),
    10: ("original", "threeD", "original_3d", "原版立封"),
}
ACTION_KEYS = ["add", "fill", "conflict", "offline_conflict", "duplicate_conflict", "note_conflict", "skip"]
OFFLINE_KEYWORDS = ["下线", "下架", "已下", "停用", "不可用"]
PREFER_ORIGINAL_KEYWORD = "优先使用原版书封"


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_title(value):
    title = clean_text(value)
    title = re.sub(r"^[《<〈\s]+|[》>〉\s]+$", "", title)
    return re.sub(r"\s+", " ", title).strip()


def resolve_status(value):
    text = clean_text(value)
    return "offline" if any(keyword in text for keyword in OFFLINE_KEYWORDS) else "active"


def resolve_preferred_version(note):
    compact = re.sub(r"\s+", "", clean_text(note))
    return "original" if PREFER_ORIGINAL_KEYWORD in compact else "auto"


def normalize_title(value):
    return re.sub(r"[《》<>〈〉【】\[\]（）()“”\"'‘’：:，,。.!！?？、·\-—_\s]", "", str(value or "").lower())


def empty_covers():
    return {
        "custom": {"flat": "", "threeD": ""},
        "original": {"flat": "", "threeD": ""},
    }


def image_extension(image):
    path = getattr(image, "path", "") or ""
    suffix = Path(path).suffix.lower()
    if suffix in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
        return ".jpg" if suffix == ".jpeg" else suffix
    fmt = (getattr(image, "format", "") or "").lower()
    if fmt in ["jpeg", "jpg"]:
        return ".jpg"
    if fmt in ["png", "webp", "gif"]:
        return f".{fmt}"
    return ".jpg"


def image_anchor(image):
    anchor = getattr(image, "anchor", None)
    if not anchor:
        return None
    try:
        return anchor._from.row + 1, anchor._from.col + 1
    except Exception:
        return None


def extract_image_bytes(image):
    image_copy = copy.copy(image)
    return image_copy._data()


def write_image(image, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(extract_image_bytes(image))


def timestamp():
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def create_run_dir(data_dir):
    runs_dir = data_dir / "merge-runs"
    base = timestamp()
    run_dir = runs_dir / base
    suffix = 1
    while run_dir.exists():
        run_dir = runs_dir / f"{base}-{suffix:02d}"
        suffix += 1
    run_dir.mkdir(parents=True)
    return run_dir


def backup_books(data_dir):
    books_path = data_dir / "books.json"
    backups_dir = data_dir / "backups"
    backups_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")[:19]
    backup_path = backups_dir / f"books-{stamp}.json"
    shutil.copy2(books_path, backup_path)
    return backup_path


def next_book_id(books):
    max_id = 0
    for book in books:
        match = re.match(r"^book_(\d+)$", str(book.get("id", "")))
        if match:
            max_id = max(max_id, int(match.group(1)))
    return f"book_{max_id + 1:03d}"


def set_cover(covers, slot_key, value):
    version, slot, _, _ = next(item for item in COVER_COLUMNS.values() if item[2] == slot_key)
    covers[version][slot] = value


def get_cover(covers, slot_key):
    version, slot, _, _ = next(item for item in COVER_COLUMNS.values() if item[2] == slot_key)
    return (covers or empty_covers()).get(version, {}).get(slot, "")


def cover_slots():
    return [value[2] for value in COVER_COLUMNS.values()]


def cover_label(slot_key):
    return next(value[3] for value in COVER_COLUMNS.values() if value[2] == slot_key)


def rel_path(path, base):
    return path.relative_to(base).as_posix()


def read_excel_rows(excel_path, run_dir, data_dir):
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path, read_only=False, data_only=True)
    sheet_warning = ""
    if SHEET_NAME in workbook.sheetnames:
        worksheet = workbook[SHEET_NAME]
    else:
        worksheet = workbook.worksheets[0]
        sheet_warning = f"未找到指定 sheet「{SHEET_NAME}」，实际使用「{worksheet.title}」。"

    images_by_cell = defaultdict(list)
    ignored_images = []
    for index, image in enumerate(getattr(worksheet, "_images", []), 1):
        anchor = image_anchor(image)
        if not anchor:
            ignored_images.append({"imageIndex": index, "reason": "missing anchor"})
            continue
        images_by_cell[anchor].append(image)

    incoming_dir = run_dir / "incoming-covers"
    rows = []
    missing_title = 0
    for row_index in range(2, worksheet.max_row + 1):
        title = clean_title(worksheet.cell(row_index, TITLE_COL).value)
        if not title:
            has_content = any(clean_text(worksheet.cell(row_index, col).value) for col in range(1, 15))
            if has_content:
                missing_title += 1
            continue

        covers = empty_covers()
        incoming = {}
        for col, (version, slot, slot_key, _) in COVER_COLUMNS.items():
            images = images_by_cell.get((row_index, col), [])
            if not images:
                continue
            image = images[0]
            filename = f"row_{row_index:04d}_{slot_key}{image_extension(image)}"
            output_path = incoming_dir / filename
            write_image(image, output_path)
            preview_path = rel_path(output_path, data_dir)
            covers[version][slot] = preview_path
            incoming[slot_key] = preview_path

        note = clean_text(worksheet.cell(row_index, STATUS_COL).value)
        rows.append({
            "row": row_index,
            "title": title,
            "note": note,
            "status": resolve_status(note),
            "preferredVersion": resolve_preferred_version(note),
            "normalizedTitle": normalize_title(title),
            "covers": covers,
            "incoming": incoming,
            "missingCover": not any(incoming.values()),
        })

    return {
        "rows": rows,
        "sheet": worksheet.title,
        "sheetWarning": sheet_warning,
        "ignoredImages": ignored_images,
        "missingTitle": missing_title,
    }


def index_books(books):
    groups = defaultdict(list)
    for book in books:
        groups[normalize_title(book.get("title", ""))].append(book)
    return groups


def choose_existing(matches):
    active = [book for book in matches if book.get("status", "active") != "offline"]
    if len(active) == 1:
        return active[0]
    if not active and len(matches) == 1:
        return matches[0]
    return None


def classify_row(row, existing_groups, duplicate_new_keys, duplicate_existing_keys):
    key = row["normalizedTitle"]
    new_slots = [slot for slot in cover_slots() if get_cover(row["covers"], slot)]
    matches = existing_groups.get(key, [])

    if key in duplicate_new_keys or key in duplicate_existing_keys:
        return "duplicate_conflict", None, "资料库或新 Excel 中有重复书名，需要人工确认", [], []

    if not matches:
        if new_slots:
            return "add", None, "资料库中没有这本书，建议新增", new_slots, []
        return "skip", None, "资料库中没有这本书，但新 Excel 没有可导入封面", [], []

    existing = choose_existing(matches)
    if not existing:
        return "duplicate_conflict", None, "匹配到多条旧记录，需要人工确认", [], []

    if existing.get("status", "active") == "offline":
        return "offline_conflict", existing, "资料库中此书已下架，不能自动恢复", [], new_slots

    old_note = clean_text(existing.get("note", ""))
    new_note = clean_text(row.get("note", ""))
    if old_note and new_note and old_note != new_note:
        return "note_conflict", existing, "新旧备注都存在且不同，不能自动覆盖", [], new_slots

    fill_slots = []
    conflict_slots = []
    for slot in new_slots:
        if get_cover(existing.get("covers", {}), slot):
            conflict_slots.append(slot)
        else:
            fill_slots.append(slot)

    if conflict_slots:
        return "conflict", existing, "同一封面位新旧资料都存在，默认不能覆盖", fill_slots, conflict_slots
    if fill_slots or (new_note and not old_note):
        return "fill", existing, "旧资料缺少部分封面或备注，新 Excel 可以补充", fill_slots, []
    return "skip", existing, "没有新信息", [], []


def build_plan(excel_rows, books):
    existing_groups = index_books(books)
    existing_counts = {key: len(value) for key, value in existing_groups.items() if key}
    new_counts = Counter(row["normalizedTitle"] for row in excel_rows)
    duplicate_existing_keys = {key for key, count in existing_counts.items() if count > 1}
    duplicate_new_keys = {key for key, count in new_counts.items() if count > 1}

    plan = []
    for row in excel_rows:
        action, existing, reason, fill_slots, conflict_slots = classify_row(row, existing_groups, duplicate_new_keys, duplicate_existing_keys)
        plan.append({
            "action": action,
            "reason": reason,
            "row": row["row"],
            "newTitle": row["title"],
            "newStatus": row.get("status", "active"),
            "newNote": row.get("note", ""),
            "newPreferredVersion": row.get("preferredVersion", "auto"),
            "normalizedTitle": row["normalizedTitle"],
            "newCovers": row["covers"],
            "newCoverSlots": [slot for slot in cover_slots() if get_cover(row["covers"], slot)],
            "existingBookId": existing.get("id") if existing else "",
            "existingTitle": existing.get("title") if existing else "",
            "existingStatus": existing.get("status", "active") if existing else "",
            "existingNote": existing.get("note", "") if existing else "",
            "existingCovers": existing.get("covers", empty_covers()) if existing else empty_covers(),
            "fillSlots": fill_slots,
            "conflictSlots": conflict_slots,
        })
    return plan


def summarize(plan, excel_info):
    summary = {
        "newExcelRows": len(excel_info["rows"]),
        "add": 0,
        "fill": 0,
        "conflict": 0,
        "offlineConflict": 0,
        "duplicateConflict": 0,
        "noteConflict": 0,
        "skip": 0,
        "missingTitle": excel_info["missingTitle"],
        "missingCover": sum(1 for row in excel_info["rows"] if row["missingCover"]),
    }
    for item in plan:
        if item["action"] == "offline_conflict":
            summary["offlineConflict"] += 1
        elif item["action"] == "duplicate_conflict":
            summary["duplicateConflict"] += 1
        elif item["action"] == "note_conflict":
            summary["noteConflict"] += 1
        else:
            summary[item["action"]] += 1
    return summary


def preview_src(path, run_dir):
    if not path:
        return ""
    normalized = str(path).replace("\\", "/")
    run_prefix = f"merge-runs/{run_dir.name}/"
    if normalized.startswith(run_prefix):
        return normalized[len(run_prefix):]
    return f"../../{normalized}"


def img_cell(path, label, run_dir):
    if path:
        return f'<div class="cover-cell"><span>{html.escape(label)}</span><img src="{html.escape(preview_src(path, run_dir))}" alt="{html.escape(label)}"></div>'
    return f'<div class="cover-cell"><span>{html.escape(label)}</span><div class="missing">缺失</div></div>'


def render_covers(covers, run_dir):
    return "".join(img_cell(get_cover(covers, slot), cover_label(slot), run_dir) for slot in cover_slots())


def build_preview_html(summary, plan, run_dir):
    rows = []
    for item in plan:
        action = item["action"]
        reason = item["reason"]
        old_title = item["existingTitle"] or "无"
        new_status = "已下架" if item.get("newStatus") == "offline" else "正常"
        old_status = "已下架" if item.get("existingStatus") == "offline" else "正常" if item.get("existingTitle") else "无"
        new_note = item.get("newNote") or "无"
        old_note = item.get("existingNote") or "无"
        rows.append(f"""
        <article class="merge-card {html.escape(action)}" data-action="{html.escape(action)}">
          <header>
            <div>
              <h2>{html.escape(item["newTitle"])}</h2>
              <p>新表状态：{new_status} · 匹配旧书：{html.escape(old_title)} · 旧状态：{old_status}</p>
            </div>
            <strong>{html.escape(action)}</strong>
          </header>
          <div class="reason">{html.escape(reason)}</div>
          <div class="notes">
            <div><b>新备注</b><p>{html.escape(new_note)}</p></div>
            <div><b>旧备注</b><p>{html.escape(old_note)}</p></div>
          </div>
          <div class="compare">
            <section>
              <h3>新 Excel 封面</h3>
              <div class="covers">{render_covers(item["newCovers"], run_dir)}</div>
            </section>
            <section>
              <h3>资料库旧封面</h3>
              <div class="covers">{render_covers(item["existingCovers"], run_dir)}</div>
            </section>
          </div>
        </article>
        """)
    summary_html = html.escape(json.dumps(summary, ensure_ascii=False, indent=2))
    filters = "\n".join(f'<button data-filter="{key}">{key}</button>' for key in ["all"] + ACTION_KEYS)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Excel 合并预览</title>
  <style>
    body {{ margin: 0; padding: 24px; background: #f3efe7; color: #25221d; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif; }}
    h1 {{ margin: 0 0 12px; }}
    .summary {{ white-space: pre-wrap; background: #fffdf8; border: 1px solid #e2d8ca; border-radius: 8px; padding: 14px; margin-bottom: 12px; }}
    .filters {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 16px; }}
    button {{ border: 1px solid #e2d8ca; border-radius: 999px; background: #fffdf8; padding: 8px 12px; cursor: pointer; }}
    button.active {{ color: #116a5b; border-color: #116a5b; background: #edf8f5; }}
    .grid {{ display: grid; gap: 14px; }}
    .merge-card {{ background: #fffdf8; border: 1px solid #e2d8ca; border-left: 5px solid #8b7f70; border-radius: 8px; padding: 14px; box-shadow: 0 12px 28px rgba(78, 58, 34, 0.08); }}
    .merge-card.add {{ border-left-color: #087443; }}
    .merge-card.fill {{ border-left-color: #116a5b; }}
    .merge-card.conflict {{ border-left-color: #b7791f; }}
    .merge-card.offline_conflict {{ border-left-color: #b42318; background: #fff7f5; }}
    .merge-card.duplicate_conflict {{ border-left-color: #7f56d9; }}
    .merge-card.note_conflict {{ border-left-color: #d97706; background: #fffaf0; }}
    .merge-card.skip {{ border-left-color: #8b7f70; opacity: 0.78; }}
    .merge-card.hidden {{ display: none; }}
    header {{ display: flex; justify-content: space-between; gap: 12px; align-items: flex-start; margin-bottom: 10px; }}
    h2, h3, p {{ margin: 0; }}
    header p {{ margin-top: 5px; color: #746f66; font-size: 14px; }}
    header strong {{ border-radius: 999px; padding: 5px 9px; background: #f7f1e8; }}
    .reason {{ margin-bottom: 12px; color: #5d5449; }}
    .notes {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 12px; }}
    .notes div {{ background: #f7f1e8; border-radius: 7px; padding: 9px; }}
    .notes b {{ display: block; margin-bottom: 4px; }}
    .notes p {{ margin: 0; white-space: pre-wrap; color: #5d5449; }}
    .offline_conflict .reason {{ color: #b42318; font-weight: 700; }}
    .compare {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }}
    section {{ min-width: 0; }}
    section h3 {{ font-size: 15px; margin-bottom: 8px; }}
    .covers {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 8px; }}
    .cover-cell span {{ display: block; color: #746f66; font-size: 12px; margin-bottom: 5px; }}
    img, .missing {{ width: 100%; aspect-ratio: 2 / 3; object-fit: contain; background: #f5ecdf; border: 1px solid #e2d8ca; border-radius: 6px; }}
    .missing {{ display: grid; place-items: center; color: #9a8d7b; border-style: dashed; font-size: 12px; }}
    @media (max-width: 900px) {{ .compare {{ grid-template-columns: 1fr; }} }}
  </style>
</head>
<body>
  <h1>Excel 合并预览</h1>
  <div class="summary">{summary_html}</div>
  <div class="filters">{filters}</div>
  <main class="grid">{''.join(rows)}</main>
  <script>
    const buttons = Array.from(document.querySelectorAll("button[data-filter]"));
    const cards = Array.from(document.querySelectorAll(".merge-card"));
    buttons[0].classList.add("active");
    buttons.forEach(button => button.addEventListener("click", () => {{
      buttons.forEach(item => item.classList.toggle("active", item === button));
      const filter = button.dataset.filter;
      cards.forEach(card => card.classList.toggle("hidden", filter !== "all" && card.dataset.action !== filter));
    }}));
  </script>
</body>
</html>
"""


def copy_incoming_cover_to_data(path_value, data_dir, book_id, slot_key):
    source = data_dir / path_value
    if not source.exists():
        return ""
    ext = source.suffix or ".jpg"
    covers_dir = data_dir / "covers"
    covers_dir.mkdir(parents=True, exist_ok=True)
    target = covers_dir / f"{book_id}_{slot_key}_merge_{datetime.now().strftime('%H%M%S%f')}{ext}"
    shutil.copy2(source, target)
    return target.relative_to(data_dir).as_posix()


def apply_plan(plan, books, data_dir):
    applied = {"add": 0, "fill": 0, "skipped": 0, "backup": ""}
    applied["backup"] = str(backup_books(data_dir))
    by_id = {book["id"]: book for book in books}

    for item in plan:
        if item["action"] == "add":
            book_id = next_book_id(books)
            covers = empty_covers()
            for slot in item["newCoverSlots"]:
                copied = copy_incoming_cover_to_data(get_cover(item["newCovers"], slot), data_dir, book_id, slot)
                if copied:
                    set_cover(covers, slot, copied)
            now = datetime.now().date().isoformat()
            books.append({
                "id": book_id,
                "title": item["newTitle"],
                "note": item.get("newNote", ""),
                "status": item.get("newStatus", "active"),
                "preferredVersion": item.get("newPreferredVersion", "auto"),
                "covers": covers,
                "createdAt": now,
                "updatedAt": now,
            })
            applied["add"] += 1
        elif item["action"] == "fill":
            book = by_id.get(item["existingBookId"])
            if not book:
                applied["skipped"] += 1
                continue
            changed = False
            new_note = clean_text(item.get("newNote", ""))
            if new_note and not clean_text(book.get("note", "")):
                book["note"] = new_note
                if item.get("newPreferredVersion") == "original" and book.get("preferredVersion", "auto") == "auto":
                    book["preferredVersion"] = "original"
                changed = True
            for slot in item["fillSlots"]:
                if get_cover(book.get("covers", {}), slot):
                    continue
                copied = copy_incoming_cover_to_data(get_cover(item["newCovers"], slot), data_dir, book["id"], slot)
                if copied:
                    set_cover(book["covers"], slot, copied)
                    changed = True
            if changed:
                book["updatedAt"] = datetime.now().date().isoformat()
                applied["fill"] += 1
            else:
                applied["skipped"] += 1
        else:
            applied["skipped"] += 1

    (data_dir / "books.json").write_text(json.dumps(books, ensure_ascii=False, indent=2), encoding="utf-8")
    return applied


def run_merge(excel_path, data_dir, apply=False):
    data_dir = Path(data_dir)
    books_path = data_dir / "books.json"
    books = json.loads(books_path.read_text(encoding="utf-8")) if books_path.exists() else []
    run_dir = create_run_dir(data_dir)

    excel_info = read_excel_rows(excel_path, run_dir, data_dir)
    plan = build_plan(excel_info["rows"], books)
    summary = summarize(plan, excel_info)

    (run_dir / "merge-summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (run_dir / "merge-plan.json").write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    (run_dir / "merge-preview.html").write_text(build_preview_html(summary, plan, run_dir), encoding="utf-8")

    apply_report = None
    if apply:
        apply_report = apply_plan(plan, books, data_dir)
        (run_dir / "apply-report.json").write_text(json.dumps(apply_report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Excel rows: {summary['newExcelRows']}")
    print(f"add: {summary['add']}, fill: {summary['fill']}, conflict: {summary['conflict']}, offline_conflict: {summary['offlineConflict']}, duplicate_conflict: {summary['duplicateConflict']}, skip: {summary['skip']}")
    print(f"Report: {run_dir}")
    if apply_report:
        print(f"Applied add: {apply_report['add']}, fill: {apply_report['fill']}, skipped: {apply_report['skipped']}")


def main():
    parser = argparse.ArgumentParser(description="Create a dry-run merge plan for a new Excel cover sheet.")
    parser.add_argument("excel", help="Path to the new .xlsx file")
    parser.add_argument("--data", default="data", help="Data directory containing books.json and covers/")
    parser.add_argument("--apply", action="store_true", help="Apply safe add/fill actions after generating the plan")
    args = parser.parse_args()
    run_merge(Path(args.excel), Path(args.data), apply=args.apply)


if __name__ == "__main__":
    main()
